from django.conf import settings
import json
import os
import openpyxl
import pandas as pd
from decimal import Decimal
import requests
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib import messages
from .models import Product, WishlistItem, Cart, Wishlist, Contribution, ProductImage, Tag, ProductVariant, \
    Category, Gift, Order, CartItem
# from django.contrib.auth.models import User
from users.models import User
from .payment_functions import initialize_payment
from dotenv import load_dotenv
from .wishlist import WishlistService
from django.db.models import Q, Count
from django.utils import timezone
from .forms import ProductForm, InventoryProductForm, FeaturedAndAvailableForm, CategoryForm, BulkCategoryUploadForm, \
    BulkProductUploadForm, ProductVariantForm, GiftPaymentForm, GuestCheckoutForm, CheckoutForm


# Helper function to check if user is vendor
def is_vendor(user):
    return user.is_authenticated and (user.role == 'vendor')  # Check if user is admin or vendor


def is_manager(user):
    return user.is_authenticated and user.groups.filter(name='Managers').exists()


# Single Responsibility Principle: Separate logic for paginated product fetching.
def get_paginated_products(request, page_number=1, per_page=12):
    products = Product.objects.filter(is_active=True, available=True)
    paginator = Paginator(products, per_page)
    return paginator.get_page(page_number)


def start_payment(request):
    email = "customer@example.com"
    amount_in_ghc = 1000
    amount_in_pesewas = amount_in_ghc * 100  # Convert to pesewas

    payment_url = initialize_payment(email, amount_in_pesewas)
    if "https" in payment_url:
        # Redirect the user to the payment page
        return redirect(payment_url)
    else:
        # Handle error
        return HttpResponse(payment_url)


# ########## CONFIRM TRANSACTIONS AND LOG THEM ##############
@csrf_exempt
def paystack_webhook(request):
    if request.method == 'POST':
        # Get the request body and decode it
        event = json.loads(request.body.decode('utf-8'))

        # Retrieve the event type and details
        event_type = event.get('event', None)
        data = event.get('data', {})

        # Paystack sends different events, we are only interested in 'charge.success'
        if event_type == 'charge.success':
            # Verify the event by contacting Paystack's API
            payment_reference = data.get('reference')

            if payment_reference:
                verified = verify_paystack_payment(payment_reference)

                if verified:
                    # Retrieve contribution data from session (you may need to store it persistently in some cases)
                    contribution_data = request.session.get('contribution_data', {})
                    amount = Decimal(contribution_data.get('amount'))
                    wishlist_id = contribution_data.get('wishlist_id')
                    contributor_name = contribution_data.get('contributor_name')
                    contact_info = contribution_data.get('contact_info')
                    message = contribution_data.get('message')
                    item_id = contribution_data.get('item_id')

                    wishlist = Wishlist.objects.get(id=wishlist_id)
                    if item_id:
                        # Specific item contribution
                        item = WishlistItem.objects.get(id=item_id, wishlist=wishlist)
                        handle_item_contribution(item, amount, contributor_name, contact_info, message)
                    else:
                        # General contribution
                        handle_general_contribution(wishlist, amount, contributor_name, contact_info, message)

                    # Clear session data after successful contribution
                    del request.session['contribution_data']

                    return JsonResponse({'status': 'success', 'message': 'Payment verified and contribution logged'},
                                        status=200)
                else:
                    return JsonResponse({'status': 'error', 'message': 'Payment verification failed'}, status=400)

        return JsonResponse({'status': 'success'}, status=200)

    return JsonResponse({'status': 'invalid request'}, status=400)


def verify_paystack_payment(reference):
    """Verify the payment reference with Paystack API to confirm payment."""
    url = f'https://api.paystack.co/transaction/verify/{reference}'
    headers = {
        "Authorization": f"Bearer {os.getenv('PAY_SECRET')}"
    }

    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        response_data = response.json()
        return response_data.get('status') and response_data['data']['status'] == 'success'

    return False


# Open/Closed Principle: Extendable logic for actions based on login state.
def index(request):
    page_number = request.GET.get('page', 1)
    products_page = get_paginated_products(request, page_number)

    # Fetch the user's non-expired wishlists (if logged in)
    wishlists = Wishlist.objects.not_expired().filter(user=request.user) if request.user.is_authenticated else []

    context = {
        'products_page': products_page,
        'wishlists': wishlists  # Pass the user's wishlists to the template
    }
    return render(request, 'synergy_mall/index.html', context)


@require_POST
@login_required
def add_to_wishlist_ajax(request):
    """
    Handles the AJAX request to add a product to a selected wishlist.
    """
    import json
    data = json.loads(request.body)

    product_id = data.get('product_id')
    wishlist_id = data.get('wishlist_id')

    try:
        product = Product.objects.get(id=product_id)

        # Fetch the user's non-expired wishlists
        wishlist = Wishlist.objects.not_expired().get(id=wishlist_id, user=request.user)

        # Check if the product is already in the wishlist
        wishlist_item, created = WishlistItem.objects.get_or_create(
            wishlist=wishlist,
            product=product
        )

        if created:
            response_data = {
                'success': True,
                'product_name': product.name,
                'wishlist_title': wishlist.title
            }
        else:
            response_data = {
                'success': False,
                'message': 'Product is already in the wishlist.'
            }

    except Product.DoesNotExist:
        response_data = {
            'success': False,
            'message': 'Product not found.'
        }
    except Wishlist.DoesNotExist:
        response_data = {
            'success': False,
            'message': 'Wishlist not found, does not belong to you, or is expired.'
        }

    return JsonResponse(response_data)


@login_required
def create_wishlist(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        privacy = request.POST.get('privacy', 'private')
        expiry_date = request.POST.get('expiry_date', None)

        wishlist_service = WishlistService(request.user)

        try:
            # Attempt to create the wishlist
            wishlist = wishlist_service.create_wishlist(
                title=title,
                description=description,
                privacy=privacy,
                expiry_date=expiry_date,
            )
            messages.success(request, f'Wishlist "{wishlist.title}" created successfully!')
            return redirect('view_wishlist', wishlist_id=wishlist.id)
        except ValidationError as e:
            # Display the error message
            messages.error(request, str(e))

    return render(request, 'synergy_mall/create_wishlist.html')


@login_required
def delete_wishlist(request, wishlist_id):
    wishlist_service = WishlistService(request.user)

    try:
        wishlist_service.delete_wishlist(wishlist_id)
        messages.success(request, 'Wishlist deleted successfully!')
    except ValidationError as e:
        messages.error(request, str(e))  # Display the error message

    return redirect('my_wishlists')


@login_required
def edit_wishlist(request, wishlist_id):
    wishlist_service = WishlistService(request.user)
    wishlist = wishlist_service._get_user_wishlist(wishlist_id)

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        privacy = request.POST.get('privacy', 'private')
        expiry_date = request.POST.get('expiry_date', None)

        try:
            # Attempt to update the wishlist
            wishlist_service.update_wishlist(
                wishlist_id=wishlist.id,
                title=title,
                description=description,
                privacy=privacy,
                expiry_date=expiry_date
            )
            messages.success(request, f'Wishlist "{wishlist.title}" updated successfully!')
            return redirect('view_wishlist', wishlist_id=wishlist.id)

        except ValidationError as e:
            # Show validation error messages
            messages.error(request, str(e))

    context = {
        'wishlist': wishlist
    }
    return render(request, 'synergy_mall/edit_wishlist.html', context)


# @login_required
def view_wishlist(request, wishlist_id):
    wishlist_service = WishlistService(request.user)
    wishlist = wishlist_service.get_wishlist(wishlist_id)

    general_contributions = wishlist.contribution_set.filter(wishlist_item__isnull=True)

    user_info = {}
    if request.user.is_authenticated:
        user_info = {
            'first_name': request.user.first_name,
            'surname': request.user.surname,
            'other_names': request.user.other_names,
            'email': request.user.email,
        }

    context = {
        'wishlist': wishlist,
        'items': wishlist.ordered_items(),  # Load items in the correct order
        'general_contributions': general_contributions,  # Pass general contributions to the template
        'user_info': user_info,
    }
    return render(request, 'synergy_mall/view_wishlist.html', context)


@login_required
def update_wishlist_order(request):
    """Update the order of the wishlist items."""
    if request.method == "POST":
        data = json.loads(request.body)
        wishlist_id = data.get('wishlist_id')
        ordered_item_ids = data.get('ordered_item_ids')

        wishlist = Wishlist.objects.get(id=wishlist_id, user=request.user)
        for index, item_id in enumerate(ordered_item_ids):
            WishlistItem.objects.filter(id=item_id, wishlist=wishlist).update(ordering=index)

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request'})


@login_required
def my_wishlists(request):
    """
    View to list all the logged-in user's wishlists.
    """
    wishlist_service = WishlistService(request.user)
    wishlists = wishlist_service.get_user_wishlists()

    # Paginate the wishlists (e.g., 10 per page)
    paginator = Paginator(wishlists, 9)  # Show 9 wishlists per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj
    }

    return render(request, 'synergy_mall/my_wishlists.html', context)


def all_wishlists(request):
    # Fetch all public, non-expired wishlists
    wishlists = Wishlist.objects.public().not_expired()

    # Paginate the wishlists (e.g., 10 per page)
    paginator = Paginator(wishlists, 9)  # Show 9 wishlists per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj
    }
    return render(request, 'synergy_mall/all_wishlists.html', context)


def update_wishlist_item(request, wishlist_id, item_id):
    wishlist = get_object_or_404(Wishlist, id=wishlist_id, user=request.user)
    item = get_object_or_404(WishlistItem, id=item_id, wishlist=wishlist)

    if request.method == "POST":
        quantity = request.POST.get('quantity')
        if quantity and int(quantity) > 0:
            item.quantity = int(quantity)
            item.save()
    return redirect('view_wishlist', wishlist_id=wishlist.id)


def remove_wishlist_item(request, wishlist_id, item_id):
    wishlist = get_object_or_404(Wishlist, id=wishlist_id, user=request.user)
    item = get_object_or_404(WishlistItem, id=item_id, wishlist=wishlist)

    # Remove the item from the wishlist
    item.delete()
    return redirect('view_wishlist', wishlist_id=wishlist.id)


# ########## CONTRIBUTION VIEWS ##########

# def contribute_to_wishlist(request, wishlist_id):
#     """ This is contribution through paystack
#         Since payment confirmation is unavailable in dev mode
#         This will be disabled till w ego live on paystack"""
#
#     wishlist = get_object_or_404(Wishlist, id=wishlist_id)
#
#     try:
#         amount = Decimal(request.POST.get('amount'))  # Ensure it's passed as Decimal
#     except (TypeError, ValueError):
#         messages.error(request, "Invalid contribution amount.")
#         return redirect('view_wishlist', wishlist_id=wishlist_id)
#
#     if amount <= 0:
#         messages.error(request, "Contribution amount must be greater than zero.")
#         return redirect('view_wishlist', wishlist_id=wishlist_id)
#
#     contributor_name = request.POST.get('contributor_name')
#     contact_info = request.POST.get('contact_info')
#     message = request.POST.get('message', '')  # Optional message
#     item_id = request.POST.get('item_id')  # This will be null for general contributions
#
#     # Store contribution details in session to be used after payment confirmation
#     request.session['contribution_data'] = {
#         'wishlist_id': wishlist.id,
#         'contributor_name': contributor_name,
#         'contact_info': contact_info,
#         'message': message,
#         'amount': str(amount),
#         'item_id': item_id,  # Optional: Can be null for general contributions
#     }
#
#     # Start payment
#     email = contact_info  # Assuming the contact info is an email
#     amount_in_pesewas = int(amount * 100)  # Convert to pesewas
#
#     payment_url = initialize_payment(email, amount_in_pesewas)
#     if "https" in payment_url:
#         # Redirect the user to the payment page
#         return redirect(payment_url)
#     else:
#         # Handle error (payment initialization failed)
#         messages.error(request, "Failed to initialize payment. Please try again.")
#         return redirect('view_wishlist', wishlist_id=wishlist_id)

### This is before payment integration


# def contribute_to_wishlist(request, wishlist_id):
#     wishlist = get_object_or_404(Wishlist, id=wishlist_id)
#
#     try:
#         amount = Decimal(request.POST.get('amount'))  # Ensure it's passed as Decimal
#     except (TypeError, ValueError):
#         messages.error(request, "Invalid contribution amount.")
#         return redirect('view_wishlist', wishlist_id=wishlist_id)
#
#     if amount <= 0:
#         messages.error(request, "Contribution amount must be greater than zero.")
#         return redirect('view_wishlist', wishlist_id=wishlist_id)
#
#     contributor_name = request.POST.get('contributor_name')
#     contact_info = request.POST.get('contact_info')
#     message = request.POST.get('message', '')  # Optional message
#     item_id = request.POST.get('item_id')  # This will be null for general contributions
#
#     if item_id:
#         # Specific item contribution
#         item = get_object_or_404(WishlistItem, id=item_id, wishlist=wishlist)
#         handle_item_contribution(item, amount, contributor_name, contact_info, message)
#         messages.success(request, f"Successfully contributed ${amount} to {item.product.name}.")
#     else:
#         # General contribution, apply to the first item in the wishlist that needs funds
#         handle_general_contribution(wishlist, amount, contributor_name, contact_info, message)
#         messages.success(request, f"Successfully contributed ${amount} to {wishlist.title}.")
#
#     return redirect('view_wishlist', wishlist_id=wishlist.id)


def contribute_to_wishlist(request, wishlist_id):
    """ This contribution combines both development environment and produciton
        by checking the ENV_MODE in settings.py"""
    wishlist = get_object_or_404(Wishlist, id=wishlist_id)

    try:
        amount = Decimal(request.POST.get('amount'))
        print(f"Amount to contribute: {amount}")
    except (TypeError, ValueError):
        messages.error(request, "Invalid contribution amount.")
        return redirect('view_wishlist', wishlist_id=wishlist_id)

    if amount <= 10:
        messages.error(request, "Contribution amount must be 10 or more.")
        return redirect('view_wishlist', wishlist_id=wishlist_id)

    contributor_name = request.POST.get('contributor_name')
    contact_info = request.POST.get('contact_info')
    message = request.POST.get('message', '')
    item_id = request.POST.get('item_id')

    if settings.ENV_MODE == 'development':
        # Directly log the contribution in development mode

        if item_id:
            item = get_object_or_404(WishlistItem, id=item_id, wishlist=wishlist)
            handle_item_contribution(item, amount, contributor_name, contact_info, message)
            messages.success(request, f"Successfully contributed ${amount} to {item.product.name}.")
        else:
            handle_general_contribution(wishlist, amount, contributor_name, contact_info, message)
            messages.success(request, f"Successfully contributed ${amount} to {wishlist.title}.")

        return redirect('view_wishlist', wishlist_id=wishlist.id)

    else:
        # Initialize payment in production mode
        email = contact_info
        amount_in_pesewas = int(amount * 100)

        payment_url = initialize_payment(email, amount_in_pesewas)
        if "https" in payment_url:
            return redirect(payment_url)
        else:
            messages.error(request, "Failed to initialize payment. Please try again.")
            return redirect('view_wishlist', wishlist_id=wishlist_id)


def handle_item_contribution(item, amount, contributor_name, contact_info, message):
    """Handle specific item contribution and handle surplus if any."""

    # Log the contribution
    Contribution.objects.create(
        wishlist_item=item,
        wishlist=item.wishlist,
        contributor_name=contributor_name,
        contact_info=contact_info,
        message=message,
        amount=amount,
        date=timezone.now()
    )

    remaining = item.amount_remaining()

    if amount <= remaining:
        item.amount_paid += amount
        item.save()
    else:
        surplus = amount - remaining
        item.amount_paid += remaining
        distribute_surplus(item.wishlist, surplus)
    item.save()

    # Check if the item is fully funded and create an order
    if item.amount_paid >= item.product.price:
        Order.objects.create(
            user=item.wishlist.user,
            product=item.product,
            quantity=1,
            total_price=item.product.price,
            status='pending',
            wishlist_item=item,
        )
        messages.success(request, f"Wishlist item '{item.product.name}' has been fully"
                                  f"funded and an order has been placed.")


def handle_general_contribution(wishlist, amount, contributor_name, contact_info, message):
    """Handle general contributions and apply to items in the preferred order."""
    # Log the contribution with additional details
    Contribution.objects.create(
        wishlist=wishlist,
        contributor_name=contributor_name,
        contact_info=contact_info,
        message=message,
        amount=amount,
        date=timezone.now()
    )

    items = wishlist.ordered_items()
    for item in items:
        remaining = item.amount_remaining()
        if amount <= remaining:
            item.amount_paid += amount
            item.save()
            amount = Decimal(0)
            break
        else:   # when contribution amount is greater than remaining amount,
            item.amount_paid += remaining
            item.save()
            amount -= remaining

    # If there's any surplus after all items, add it to user's cash_on_hand
    if amount > 0:
        wishlist.user.cash += amount
        wishlist.user.save()


def distribute_surplus(wishlist, surplus):
    """Distribute surplus funds to other items in the wishlist."""
    items = wishlist.ordered_items()
    for item in items:
        if item.amount_remaining() > 0:
            remaining = item.amount_remaining()
            if surplus <= remaining:
                item.amount_paid += surplus
                item.save()
                return
            else:
                surplus -= remaining
                item.amount_paid += remaining
                item.save()

    # If any surplus is left after all items are funded, add to user's cash_on_hand
    if surplus > 0:
        wishlist.user.cash += surplus
        wishlist.user.save()


# ########## PRODUCT VIEWS ##########


def search_product(request):
    """
    Search products and their variants based on query terms.
    - Rank results by relevance: Product title and description matches are prioritized.
    """
    query = request.GET.get('q', '').strip()  # Get search terms from query parameter
    products = Product.objects.none()  # Default empty queryset
    wishlists = Wishlist.objects.not_expired().filter(user=request.user) if request.user.is_authenticated else []

    if query:
        search_terms = query.split()

        title_q = Q()
        description_q = Q()
        other_q = Q()

        # Build search conditions for each term
        for term in search_terms:
            title_q |= Q(name__icontains=term)
            description_q |= Q(description__icontains=term)
            other_q |= (
                Q(category__name__icontains=term) |
                Q(tags__name__icontains=term) |
                Q(variants__color__icontains=term) |
                Q(variants__size__icontains=term)
            )

        # Filter the products with relevance-based annotation
        products = Product.objects.filter(
            title_q | description_q | other_q
        ).distinct().annotate(
            relevance=(
                Count('name', filter=title_q) * 3 +
                Count('description', filter=description_q) * 2 +
                Count('category', filter=other_q)
            )
        ).order_by('-relevance', 'name')

    context = {
        'products': products,
        'query': query,
        'wishlists': wishlists,
    }
    return render(request, 'synergy_mall/search_results.html', context)


@user_passes_test(is_vendor)  # Restrict access to vendors only
def add_product(request):
    if request.method == 'POST':
        product_form = ProductForm(request.POST, request.FILES)

        if product_form.is_valid():
            # Save the product but don't commit yet (because we need to assign the vendor)
            product = product_form.save(commit=False)
            product.vendor = request.user  # Assign the current user as the vendor
            product.save()

            # Handle multiple images
            images = request.FILES.getlist('images')
            for image in images:
                ProductImage.objects.create(product=product, image=image)

            # Handle product variants (colors and sizes)
            # colors = product_form.cleaned_data['colors']  # This will be a list of colors (or empty)
            # sizes = product_form.cleaned_data['sizes']  # This will be a list of sizes (or empty)
            colors = request.POST.getlist('colors')
            sizes = request.POST.getlist('sizes')
            print(colors)

            if colors or sizes:  # Create variants only if colors or sizes are provided
                for color in colors or [None]:  # If no colors, use None
                    for size in sizes or [None]:  # If no sizes, use None
                        ProductVariant.objects.create(product=product, color=color, size=size)

            messages.success(request, 'Product and variants added successfully!')
            return redirect('product_list')
    else:
        product_form = ProductForm()

    return render(request, 'synergy_mall/add_product.html', {'form': product_form})


@user_passes_test(is_vendor)
def bulk_add_products(request):
    if request.method == 'POST':
        form = BulkProductUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            try:
                # Load the Excel file using pandas
                df = pd.read_excel(excel_file, engine='openpyxl')

                # Validate and process each row
                for index, row in df.iterrows():
                    product_name = row['Name']
                    description = row['Description']
                    category_name = row['Category']
                    price = row['Price']
                    sale_price = row.get('Sale Price', None)  # Sale price is optional
                    color = row.get('Color', None)
                    size = row.get('Size', None)
                    sku = row['SKU']
                    stock = row['Stock']

                    # Get or create the category
                    category, created = Category.objects.get_or_create(name=category_name)

                    # Create the product (if it doesn't already exist)
                    product, created = Product.objects.get_or_create(
                        name=product_name,
                        description=description,
                        price=price,
                        sale_price=sale_price,
                        vendor=request.user,  # Assign the current user as the vendor
                        category=category,
                    )

                    # Create the variant (color and size can be None if not specified)
                    ProductVariant.objects.create(
                        product=product,
                        color=color,
                        size=size,
                        sku=sku,
                        stock=stock,
                    )

                messages.success(request, 'Products added successfully!')
                return redirect('product_list')

            except Exception as e:
                messages.error(request, f'Error processing file: {e}')
    else:
        form = BulkProductUploadForm()

    return render(request, 'synergy_mall/bulk_add_products.html', {'form': form})


def download_bulk_product_template(request):
    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Template"

    # Define the headers (columns) for the Excel file
    headers = [
        'Name', 'Description', 'Category', 'Price', 'Sale Price (Optional)', 'Color (Optional)',
        'Size (Optional)', 'SKU', 'Stock'
    ]

    # Add the headers to the first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Get all available categories from the database
    categories = Category.objects.values_list('name', flat=True)

    # Create a separate sheet for categories
    category_sheet = wb.create_sheet(title="Categories")
    for index, category in enumerate(categories, 1):
        category_sheet.cell(row=index, column=1, value=category)

    # Define the data validation for category dropdown
    category_range = f'Categories!$A$1:$A${len(categories)}'  # Adjust the range based on the number of categories
    category_validation = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=category_range, allow_blank=False
    )

    # Apply the validation to the category column (column 3)
    category_validation.add(f'C2:C1048576')  # Excel's max row limit for validation

    ws.add_data_validation(category_validation)

    # Set the response to return the Excel file as an attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Define the filename for the Excel file
    filename = f'bulk_product_template_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save the workbook to the response
    wb.save(response)

    return response


@user_passes_test(is_vendor)
def upload_product_images(request, product_id):
    product = get_object_or_404(Product, id=product_id, vendor=request.user)

    if request.method == 'POST':
        images = request.FILES.getlist('images')
        for image in images:
            ProductImage.objects.create(product=product, image=image)
        messages.success(request, 'Images uploaded successfully!')
        return redirect('product_list')

    return render(request, 'synergy_mall/upload_images.html', {'product': product})


def view_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    variants = product.variants.all()

    # Group variants by available colors and sizes
    available_colors = variants.values_list('color', flat=True).distinct()
    available_sizes = variants.values_list('size', flat=True).distinct()

    context = {
        'product': product,
        'variants': variants,
        'available_colors': available_colors,
        'available_sizes': available_sizes,
    }
    return render(request, 'synergy_mall/view_product.html', context)


# @user_passes_test(is_vendor)
# def edit_product(request, product_id):
#     product = get_object_or_404(Product, id=product_id, vendor=request.user)
#
#     if request.method == 'POST':
#         form = InventoryProductForm(request.POST, request.FILES, instance=product)
#         if form.is_valid():
#             product = form.save(commit=False)
#             product.save()
#
#             # Process the tags input
#             tags_input = request.POST.get('tags')
#             if tags_input:
#                 tag_names = [tag.strip() for tag in tags_input.split(',')]
#                 product.tags.set(Tag.objects.filter(name__in=tag_names))
#
#             # Update the variants (colors, sizes) if provided
#             colors = form.cleaned_data['colors']  # List of colors (can be empty)
#             sizes = form.cleaned_data['sizes']  # List of sizes (can be empty)
#
#             # Clear existing variants and recreate only if colors or sizes are provided
#             product.variants.all().delete()
#
#             if colors or sizes:
#                 for color in colors:
#                     for size in sizes:
#                         ProductVariant.objects.create(product=product, color=color, size=size)
#
#             messages.success(request, 'Product and variants updated successfully!')
#             return redirect('product_list')
#     else:
#         form = InventoryProductForm(instance=product)
#
#     return render(request, 'synergy_mall/edit_product.html', {'form': form, 'product': product})

@user_passes_test(is_vendor)
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id, vendor=request.user)

    if request.method == 'POST':
        form = InventoryProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product = form.save(commit=False)
            product.save()

            # Process tags
            tags_input = request.POST.get('tags')
            if tags_input:
                tag_names = [tag.strip() for tag in tags_input.split(',')]
                product.tags.set(Tag.objects.filter(name__in=tag_names))

            # Handle images (if any)
            images = request.FILES.getlist('images')
            for image in images:
                ProductImage.objects.create(product=product, image=image)

            messages.success(request, 'Product details updated successfully!')
            return redirect('product_list')
    else:
        form = InventoryProductForm(instance=product)

    return render(request, 'synergy_mall/edit_product.html', {'form': form, 'product': product})


@user_passes_test(is_vendor)
def remove_product_image(request, product_id, image_id):
    product = get_object_or_404(Product, id=product_id, vendor=request.user)
    image = get_object_or_404(ProductImage, id=image_id, product=product)
    image.delete()
    messages.success(request, 'Image removed successfully.')
    return redirect('edit_product', product_id=product_id)


def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    variants = product.variants.all()

    context = {
        'product': product,
        'vendor_profile': product.vendor.vendor_profile,
        'images': product.images.all(),
        'variants': variants,
    }
    return render(request, 'synergy_mall/product_detail.html', context)


@user_passes_test(is_vendor)
def manage_product_variants(request, product_id):
    product = get_object_or_404(Product, id=product_id, vendor=request.user)

    if request.method == 'POST':
        # Check if user is updating existing variant or creating a new one
        if 'new_color' in request.POST:
            # Handle new variant creation
            new_color = request.POST.get('new_color')
            new_size = request.POST.get('new_size')
            new_stock = request.POST.get('new_stock')
            new_price = request.POST.get('new_price')
            new_sale_price = request.POST.get('new_sale_price')

            # Automatically generate SKU for the new variant
            new_sku = f"{product.name[:3].upper()}-{new_color[:3].upper()}-{new_size[:2].upper()}"

            # Create new variant
            ProductVariant.objects.create(
                product=product,
                color=new_color,
                size=new_size,
                sku=new_sku,
                stock=new_stock,
                price=new_price,
                sale_price=new_sale_price
            )
            messages.success(request, 'New product variant created successfully!')
        else:
            # Handle existing variant update
            form = ProductVariantForm(request.POST, product=product)
            if form.is_valid():
                color = form.cleaned_data['color']
                size = form.cleaned_data['size']
                stock = form.cleaned_data['stock']
                price = form.cleaned_data['price']
                sale_price = form.cleaned_data['sale_price']

                # Check if a variant with the same color and size already exists
                variant, created = ProductVariant.objects.get_or_create(
                    product=product,
                    color=color,
                    size=size
                )

                # If the variant already exists, update the stock and price (not SKU)
                variant.stock = stock
                variant.price = price
                variant.sale_price = sale_price
                variant.save()

                messages.success(request, 'Product variant updated successfully!')

        return redirect('manage_product_variants', product_id=product.id)

    # Prepopulate existing variants for dropdown
    form = ProductVariantForm(product=product)
    variants = product.variants.all()

    return render(request, 'synergy_mall/manage_product_variants.html', {
        'form': form,
        'product': product,
        'variants': variants,
    })


def vendor_products(request, vendor_id):
    vendor = get_object_or_404(User, id=vendor_id, role='vendor')  # Ensure the user is a vendor

    products = Product.objects.filter(vendor=vendor)  # Fetch all products belonging to the vendor

    context = {
        'vendor': vendor,
        'products': products,
    }

    return render(request, 'synergy_mall/vendor_products.html', context)


@user_passes_test(is_manager)
def edit_featured_and_available(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = FeaturedAndAvailableForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product featured and availability status updated!')
            return redirect('manager_dashboard')  # Redirect to manager dashboard
    else:
        form = FeaturedAndAvailableForm(instance=product)

    return render(request, 'synergy_mall/edit_featured_and_available.html', {'form': form, 'product': product})


# CATEGORIES VIEW
# @user_passes_test(is_manager)
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')  # Redirect to a page showing the list of categories
    else:
        form = CategoryForm()

    return render(request, 'synergy_mall/add_category.html', {'form': form})


def category_list(request):
    categories = Category.objects.all()  # Retrieve all categories from the database
    return render(request, 'synergy_mall/category_list.html', {'categories': categories})


def bulk_category_upload(request):
    if request.method == 'POST':
        form = BulkCategoryUploadForm(request.POST, request.FILES)

        if form.is_valid():
            # Get the uploaded file
            uploaded_file = request.FILES['file']

            # Read the file depending on its extension
            try:
                if uploaded_file.name.endswith('.xlsx'):
                    df = pd.read_excel(uploaded_file)
                elif uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    messages.error(request, 'Unsupported file format. Please upload an Excel or CSV file.')
                    return redirect('bulk_category_upload')

                # Loop through the DataFrame and create categories
                for index, row in df.iterrows():
                    category_name = row.get('name')
                    category_description = row.get('description', '')

                    if category_name:  # Ensure the name is not empty
                        Category.objects.get_or_create(
                            name=category_name,
                            defaults={'description': category_description}
                        )

                messages.success(request, 'Categories uploaded successfully!')
                return redirect('category_list')  # Redirect to the category list page
            except Exception as e:
                messages.error(request, f'Error processing file: {e}')
                return redirect('bulk_category_upload')
    else:
        form = BulkCategoryUploadForm()

    return render(request, 'synergy_mall/bulk_category_upload.html', {'form': form})


@user_passes_test(is_vendor)
def product_list(request):
    # Fetch the products for the logged-in vendor
    vendor = request.user
    products = Product.objects.filter(vendor=vendor)
    return render(request, 'synergy_mall/product_list.html', {'products': products})


@user_passes_test(is_vendor)
def suspend_product(request, product_id):
    product = get_object_or_404(Product, id=product_id, vendor=request.user)
    product.available = False
    product.save()
    messages.success(request, 'Product suspended successfully.')
    return redirect('product_list')


@user_passes_test(is_vendor)
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id, vendor=request.user)

    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully.')
        return redirect('product_list')

    return render(request, 'synergy_mall/delete_product.html', {'product': product})


# ############# CART & ORDERS ################

def view_cart(request):
    # Retrieve the cart based on user or session
    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user)
    else:
        session_key = request.session.session_key or request.session.create()
        cart, created = Cart.objects.get_or_create(session_key=session_key)

    # Calculate total price
    total_price = cart.get_total_price()

    return render(request, 'synergy_mall/cart.html', {
        'cart': cart,
        'total_price': total_price,
    })


def get_cart_items(request):
    # Retrieve cart items based on user or session
    if request.user.is_authenticated:
        cart = Cart.objects.filter(user=request.user).first()
    else:
        session_key = request.session.session_key or request.session.create()
        cart = Cart.objects.filter(session_key=session_key).first()

    if not cart:
        return JsonResponse({"success": False, "message": "Cart is empty.", "cart_items": []})

    # Serialize cart items
    cart_items = [
        {
            "id": item.id,  # Ensure item ID is included
            "product_name": item.product.name,
            "quantity": item.quantity,
            "price": float(item.price),
            "total_price": float(item.get_total_price())
        }
        for item in cart.items.all()
    ]

    return JsonResponse({"success": True, "cart_items": cart_items, "total_price": cart.get_total_price()})


def get_cart(request):
    """Retrieve or create the cart for the current user or session."""
    if request.user.is_authenticated:
        # Fetch or create a cart for the authenticated user
        cart, created = Cart.objects.get_or_create(user=request.user)
    else:
        # Use session_key for unauthenticated users
        session_key = request.session.session_key
        if not session_key:
            # Create a session if one does not exist
            request.session.create()
            session_key = request.session.session_key
        # Fetch or create a cart for the session
        cart, created = Cart.objects.get_or_create(session_key=session_key)

    return cart


def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart = get_cart(request)  # Get the appropriate cart for the user or guest

    # Add or update the cart item
    cart_item, created = CartItem.objects.get_or_create(
        cart=cart, product=product, defaults={"price": product.price, "quantity": 1}
    )
    if not created:
        cart_item.quantity += 1
        cart_item.save()

    # Return JSON response
    return JsonResponse({
        "success": True,
        "message": f"{product.name} was added to your cart.",
        "cart_item_count": cart.items.count(),
    })


def remove_from_cart(request, item_id):
    """
    Remove an item from the cart and return a JSON response.
    """
    # Determine the cart (logged-in user or session-based)
    if request.user.is_authenticated:
        cart = Cart.objects.filter(user=request.user).first()
    else:
        session_key = request.session.session_key or request.session.create()
        cart = Cart.objects.filter(session_key=session_key).first()

    # If no cart exists, return an error
    if not cart:
        return JsonResponse({"success": False, "message": "Cart not found."})

    # Attempt to find and delete the cart item
    try:
        cart_item = get_object_or_404(CartItem, id=item_id, cart=cart)
        cart_item.delete()

        # Calculate the updated total price and cart item count
        total_price = cart.get_total_price()
        cart_item_count = cart.items.count()

        return JsonResponse({
            "success": True,
            "message": "Item removed from cart.",
            "total_price": total_price,
            "cart_item_count": cart_item_count,
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


def create_order(request):
    if request.method == 'POST':
        # Collect common fields
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        shipping_address = request.POST.get('shipping_address')

        # Validate required fields
        if not all([name, email, phone_number, shipping_address]):
            messages.error(request, "All fields are required.")
            return render(request, 'synergy_mall/checkout.html', {
                'cart': get_cart(request),
                'total_price': get_cart(request).get_total_price(),
            })

        # Retrieve the cart
        cart = get_cart(request)
        if not cart or not cart.cartitem_set.exists():
            messages.error(request, "Your cart is empty.")
            return redirect('view_cart')

        # Create orders for each item in the cart
        cart_items = cart.cartitem_set.all()
        for item in cart_items:
            Order.objects.create(
                user=request.user if request.user.is_authenticated else None,
                product=item.product,
                quantity=item.quantity,
                total_price=item.get_total_price(),
                status='pending',
                shipping_address=shipping_address,
                phone_number=phone_number,
                email=email,
            )

        # Send confirmation email
        order_summary = "\n".join([f"{item.quantity} x {item.product.name} - ${item.get_total_price()}" for item in cart_items])
        total_price = cart.get_total_price()
        message = (
            f"Thank you for your order, {name}!\n\n"
            f"Order Summary:\n{order_summary}\nTotal: ${total_price}\n\n"
            f"Your order will be processed shortly.\n"
            f"Shipping to: {shipping_address}\n\n"
            f"Best regards,\nSynergy Mall Team"
        )
        send_mail(
            'Order Confirmation',
            message,
            'no-reply@example.com',
            [email],
            fail_silently=False,
        )

        # Clear the cart
        cart.cartitem_set.all().delete()

        messages.success(request, "Your order has been placed successfully!")
        return redirect('checkout_success')

    # Render the checkout page
    return render(request, 'synergy_mall/checkout.html', {
        'cart': get_cart(request),
        'total_price': get_cart(request).get_total_price(),
    })


from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse

def track_order(request):
    if request.method == "POST":
        order_number = request.POST.get('order_number')
        phone_number = request.POST.get('phone_number')

        # Validate the input
        if not (order_number and phone_number):
            return render(request, 'synergy_mall/track_order.html', {'error': 'Both fields are required.'})

        try:
            order = Order.objects.get(order_number=order_number, phone_number=phone_number)
        except Order.DoesNotExist:
            return render(request, 'synergy_mall/track_order.html', {'error': 'Order not found.'})

        # Render the order details
        return render(request, 'synergy_mall/track_order_result.html', {'order': order})

    return render(request, 'synergy_mall/track_order.html')


def buy(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        form = CheckoutForm(request.POST)
        if form.is_valid():
            # Process order and save it
            order = Order.objects.create(
                user=request.user if request.user.is_authenticated else None,
                product=product,
                quantity=1,  # Outright buy is always one item
                total_price=product.price,
                status='pending',  # Default status for a new order
                **form.cleaned_data  # Include shipping or payment details from form
            )
            # Redirect to a success page or payment processor
            return redirect('order_success')  # Replace with your success page URL
    else:
        form = CheckoutForm()

    return render(request, 'synergy_mall/checkout.html', {
        'form': form,
        'product': product,
    })


@login_required
def view_orders(request):
    orders = Order.objects.filter(user=request.user).order_by('-order_date')
    return render(request, 'synergy_mall/orders.html', {'orders': orders})


@login_required
def vendor_orders(request):
    if not request.user.role == 'vendor':
        return redirect('index')  # Redirect non-vendors to the home page

    # Fetch orders for products owned by the vendor
    vendor_products = Product.objects.filter(vendor=request.user)
    orders = Order.objects.filter(product__in=vendor_products).order_by('-order_date')

    context = {
        'orders': orders,
    }
    return render(request, 'synergy_mall/vendor_orders.html', context)


@login_required
def user_orders(request):
    # Fetch orders for the logged-in user
    orders = Order.objects.filter(user=request.user).order_by('-order_date')

    context = {
        'orders': orders,
    }
    return render(request, 'synergy_mall/user_orders.html', context)


def checkout(request):
    cart = get_cart(request)

    if request.method == "POST":
        form = CheckoutForm(request.POST)
        if form.is_valid():
            # Create orders for each cart item
            for item in cart.items.all():
                Order.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    product=item.product,
                    quantity=item.quantity,
                    total_price=item.get_total_price(),
                    status="pending",
                )
            # Clear the cart
            cart.items.all().delete()
            return redirect("order_success")
    else:
        form = CheckoutForm()

    return render(request, "synergy_mall/checkout.html", {"cart": cart, "form": form})


def checkout_success(request):
    return render(request, 'synergy_mall/checkout_success.html')


# ############# GIFTING AN ITEM ##############

@csrf_exempt
def fetch_receiver_wishlists(request):
    if request.method == "POST":
        try:
            # Parse request body
            data = json.loads(request.body)
            receiver_details = data.get('receiver_details')

            # Validate receiver details
            if not receiver_details:
                return JsonResponse({'success': False, 'message': 'Receiver details are required'}, status=400)

            # Search for receiver by email, phone, or username
            receiver = User.objects.filter(
                Q(email=receiver_details) |
                Q(phone_number=receiver_details) |
                Q(username=receiver_details)
            ).first()

            if not receiver:
                return JsonResponse({'success': False, 'message': 'User not found'}, status=404)

            # Get wishlists for the receiver
            wishlists = Wishlist.objects.filter(user=receiver).values('id', 'title')

            # Return response
            return JsonResponse({
                'success': True,
                'receiver_id': receiver.id,
                'receiver_name': f"{receiver.first_name} {receiver.surname}",
                'wishlists': list(wishlists)
            })

        except Exception as e:
            # Log the error for debugging
            print(f"Error during fetch_receiver_wishlists: {e}")
            return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'}, status=500)

    # Invalid request method
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)


def process_gift_payment(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = GiftPaymentForm(request.POST)
        receiver_id = request.POST.get('receiver_id')
        wishlist_id = request.POST.get('wishlist_id')

        if form.is_valid():
            message_to_receiver = form.cleaned_data['message_to_receiver']
            giver_contact = None

            # Get receiver and validate
            receiver = get_object_or_404(User, id=receiver_id)

            # Extract and validate `amount_given`
            try:
                amount_given = Decimal(request.POST.get('amount_given', '0'))
                if amount_given < product.price:
                    return JsonResponse({
                        'success': False,
                        'message': 'Amount must be at least the price of the product.'
                    }, status=400)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': 'Amount must be a valid number.'
                }, status=400)

            # Determine the `giver` and their contact info
            if request.user.is_authenticated:
                giver = request.user
                giver_contact = giver.phone_number
            else:
                giver = None
                giver_contact = form.cleaned_data['giver_contact']
                if not giver_contact:
                    return JsonResponse({
                        'success': False,
                        'message': 'Contact information is required for unauthenticated users.'
                    }, status=400)

            # Get or validate the wishlist
            wishlist = Wishlist.objects.filter(
                user=receiver, title="General List"
            ).first() if not wishlist_id else get_object_or_404(Wishlist, id=wishlist_id)

            # Handle surplus amount
            surplus = amount_given - product.price
            if surplus > 0:
                receiver.cash += surplus
                receiver.save()

            # Create the Gift object
            gift = Gift.objects.create(
                giver=giver,
                receiver=receiver,
                product=product,
                amount_given=amount_given,
                wishlist=wishlist,
                giver_contact=giver_contact,
                message_to_receiver=message_to_receiver,
            )

            # Add the product to the wishlist as a WishlistItem
            wishlist_item = WishlistItem.objects.create(
                wishlist=wishlist,
                product=product,
                giver_contact=giver_contact,
                message=message_to_receiver,
                amount_paid=product.price,  # Fully paid for gifts
            )

            # Create an order for the gifted item
            Order.objects.create(
                user=receiver,
                product=product,
                quantity=1,
                total_price=product.price,
                status='pending',
            )

            messages.success(request, "Gift sent successfully, and the order has been placed!")
            return redirect('index')  # Or redirect to a success page
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = GiftPaymentForm()

    return render(request, 'synergy_mall/gift_payment.html', {'form': form, 'product': product})


@login_required
def received_gifts(request):
    # Fetch all gifts for the logged-in user
    gifts = Gift.objects.filter(receiver=request.user).select_related('giver', 'product', 'wishlist')

    return render(request, 'synergy_mall/received_gifts.html', {'gifts': gifts})
